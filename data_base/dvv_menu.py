from tkinter import *
import sqlite3
from tkinter.ttk import Combobox
con = sqlite3.connect("students.db")
cur = con.cursor()
m_window = Tk()
add=[]                                                          # пустой список для добавления элементов
list_of_faks=['fak_f','fak_e','fak_m']                       # список факультетов

list_of_fak_groups=['f_1','f_2','f_3']
list_of_e_groups=['e_1','e_2','e_3']
list_of_m_groups=['m_1','m_2']                                #  список групп

list_of_all_groups = ['f_1','f_2','f_3','e_1','e_2','e_3','m_1','m_2']


###################################    очистка экрана
def clean_window():                      # очистка экрана
    for i in range(10,1000,5):
        l=Label(m_window, text="                                                                                    "
                               "                                                                                    "
                               "                                                                                    "
                               "                                                                                    ").place(x=20, y=i)
    return
###################################
def create_lists():
    cur.execute("""SELECT fak_id FROM politeh """)
    for item in cur:
        list_of_faks.append(item)

    cur.execute("""SELECT group_id FROM groups """)
    for item in cur:
        list_of_all_groups.append(item)

    cur.execute("""SELECT group_id FROM groups WHERE fak_name=
                (SELECT fak_name FROM politeh WHERE fak_id=?)""",list_of_faks[0])
    for item in cur:
        list_of_fak_groups.append(item)

    cur.execute("""SELECT group_id FROM groups WHERE fak_name=
                (SELECT fak_name FROM politeh WHERE fak_id=?)""",list_of_faks[1])
    for item in cur:
        list_of_e_groups.append(item)

    cur.execute("""SELECT group_id FROM groups WHERE fak_name=
                (SELECT fak_name FROM politeh WHERE fak_id=?)""",list_of_faks[2])
    for item in cur:
        list_of_m_groups.append(str(item))


        ###################################    в меню выбрано "Стандартная база студентов"
def choose_standart_tables():
    a,b,c=data_for_standart_tables()
    load_tables(a,b,c)
###################################    в меню выбрано "Загрузить базу студентов"
def choose_loaded_tables():
    a,b,c=data_for_loaded_database()
    load_tables(a,b,c)
###################################    создаются данные для стандартной базы
def data_for_standart_tables():
    fakultets_list, group_list, students_list = [], [], []
    fakultets_list = [('fak_f','ФАВТ','Козловский'),
                      ('fak_e', 'Экономический','Петровский'),
                      ('fak_m','Машиностроительный','Семановский')]
    group_list = [('f_1', 'Программирование','ФАВТ',0),
                  ('f_2', 'Комп. сети','ФАВТ', 0),
                  ('f_3', 'Безопастность','ФАВТ',0),
                                      ('e_1', 'Бухгалтерия', 'Экономический', 0),
                                      ('e_2', 'Внешняя экономика', 'Экономический', 0),
                                      ('e_3', 'Торговля', 'Экономический', 0),
                  ('m_1', 'Автомобилестроение','Машиностроительный', 0),
                  ('m_2', 'Сельхоз техника','Машиностроительный', 0)]
    students_list = [('st_1', 'Иванов', 'Иван', 'fak_f', 'f_1', '98'),
                     ('st_2', 'Петров', 'Петр', 'fak_f', 'f_1', '75'),
                     ('st_3', 'Сидоров', 'Сеня', 'fak_e', 'e_1', '93'),
                     ('st_4', 'Семенов', 'Ваня ', 'fak_e', 'e_1', '65'),
                     ('st_5', 'Васильков', 'Вася', 'fak_m', 'm_1', '88'),
                     ('st_6', 'Соколов', 'Федя', 'fak_f', 'f_2', '78')]
    return fakultets_list,group_list,students_list
###################################    создаются данные для загруженной базы
def data_for_loaded_database():
    fakultets_list, group_list, students_list = [], [], []
    from tkinter.filedialog import askopenfilename   #модуль для открытия файлов в диалоговом окне
    from openpyxl import load_workbook               # модуль для работы с .xlsx(excel) файлами
    file = askopenfilename()                        # создается окно для открытия файла
    wb = load_workbook(file)                         # в wb(workbook) загружаем файл(книга)
    fin_list=[group_list,fakultets_list,students_list]
    y=0
    for sheetname in wb.sheetnames:                 # для каждого листа из книги...
        a = list()                                  # список для ОДНОГО студента
        ws = wb[sheetname]                          # ws= лист (напр.:"students_excel") из книги "polit"
        all_rows = list(ws.rows)
        for i in range(1,len(all_rows)):            # чередуем строки из листа
            for cell in all_rows[i]:                # чередуем  ячейки каждой строки
                a.append(cell.value)                # достаем ЗНАЧЕНИЕ каждлй ячейки и вставляем в список
            fin_list[y].append(a)                   # список(кортеж) всех студентов
            a = list()
        y+=1
    return fakultets_list,group_list,students_list
###################################    непосредственно загрузка данных(списков) в базу данных
def load_tables(a,b,c):
    fakultets_list, group_list, students_list=a,b,c
    cur.execute("""PRAGMA foreign_keys = ON""")
    cur.execute("""DROP TABLE IF EXISTS students""")
    cur.execute("""DROP TABLE IF EXISTS politeh""")
    cur.execute("""DROP TABLE IF EXISTS groups""")

    cur.execute("""CREATE TABLE IF NOT EXISTS politeh(
               fak_id TEXT ,fak_name TEXT,dekan TEXT)""")
    cur.executemany("""INSERT INTO politeh VALUES(?,?,?)""", fakultets_list)
    ##########################
    cur.execute("""CREATE TABLE IF NOT EXISTS groups(
               group_id TEXT PRIMARY KEY,spec_name TEXT,fak_name TEXT,number_of_students TEXT NULL)""")
    cur.executemany("""INSERT INTO groups VALUES(?,?,?,?)""", group_list)
    #########################
    cur.execute("""CREATE TABLE IF NOT EXISTS students (
               st_id TEXT PRIMARY KEY,surname TEXT,name TEXT,fak_id TEXT,group_id TEXT,score TEXT,
               FOREIGN KEY (group_id) REFERENCES groups (group_id)
                                ON DELETE RESTRICT ON UPDATE CASCADE)""")
    cur.executemany("""INSERT INTO students VALUES(?,?,?,?,?,?)""", students_list)

    count_students_in_groups()
    con.commit()
    return
###################################    подсчет студентов в каждой группе и запись рез-та в таблицу GROUPS
def count_students_in_groups():
    for i in range(len(list_of_all_groups)):
        cur.execute("""SELECT COUNT(*) FROM students WHERE group_id=?""",[list_of_all_groups[i]])
        a=cur.fetchone()[0]
        b=list_of_all_groups[i]
        list=[a,b]
        cur.execute("""UPDATE groups SET number_of_students=? WHERE group_id=?""",list)
    con.commit()
    return
###################################    ПРОСМОТР ТРЕХ ТАБЛИЦ
def view_students():     ############# просмотр всех студентов-----база STUDENTS
    clean_window()
    cur.execute("""SELECT st_id,surname,name,fak_id ,group_id ,score FROM students """)
    head=["ID студента"," Фамилия ","    Имя    "," ID фак-та "," ID группы ","Средний бал"]
    b = 20
    for i in range(6):
        main_lbl = Label(m_window, text=head[i], font="Arial 12").place(x=b, y=50)
        b += 120
    a = 100
    for item in cur:
        b=20
        for i in range(6):
            main_lbl=Label(m_window, text=item[i],font="Arial 14").place(x=b, y=a)
            b+=120
        a += 30
    return

def view_politeh():#################    просмотр факультетов  --- база POLITEH
    clean_window()
    cur.execute("""SELECT fak_id,fak_name,dekan FROM politeh """)
    Label(m_window,text="факультеты ПОЛИТЕХА",font="Arial 18").place(x=160, y=10)
    head = ["ID", "Факультет", " Декан  "]
    b = 20
    for i in range(3):
        Label(m_window, text=head[i], font="Arial 12").place(x=b, y=70)
        b += 240
    a = 150
    for item in cur:
        b = 20
        for i in range(3):
            Label(m_window, text=item[i], font="Arial 12").place(x=b, y=a)
            b += 240
        a += 30
    return

def view_groups():#################    просмотр групп--таблица GROUPS
    clean_window()
    count_students_in_groups()
    cur.execute("""SELECT group_id ,spec_name ,fak_name TEXT,number_of_students FROM groups""")
    head = ["  ID группы  ", "Специальность","Факультет", "Кол-во студентов"]
    b = 20
    for i in range(4):
        Label(m_window, text=head[i], font="Arial 12").place(x = b, y = 50)
        b += 230
    a = 150
    for item in cur:
        b = 20
        for i in range(4):
            Label(m_window, text=item[i], font="Arial 12").place(x=b, y=a)
            b += 230
        a += 30
    return
###################################
def new_student():
 #   list_groups = []
    add_window=Toplevel()                       # создание дополнительного(дочернего) окна
    add_window.title("Добавление нового студента")  # заголовок окна
    add_window.geometry("370x280+700+500")
    def otmena_click():                         # закрытие окна ADD STRING без сохранения данных
        add_window.after(3, lambda: add_window.destroy())

    def add_click():                            # кнопка в окне ADD STRING--"добавления новой строки"
        for i in range(4):
            add.append(message[i].get())
        add.append(combo_fak_id.get())          #  пятый элемент-- из COMBOBOX
        add.append(combo_group_id.get())         #  шестой элемент-- из COMBOBOX
        cur = con.cursor()
        cur.execute("""INSERT INTO students (st_id ,surname ,name ,score,fak_id,group_id) VALUES (?,?,?,?,?,?)""", add)
        con.commit()
        add_window.after(3, lambda: add_window.destroy())  # закрытие окна через 3 млсек

    line = ["Введите ID","Введите фамилию: ","Введите имя:","Средний бал","Выберите фак-т:","Выберите группу:"]
    a=10
    for i in range(6):
        x = line[i]
        Label(add_window, text=x).place(x=20, y=a)              # неактивна надпись слева
        a+=30
    message1 = StringVar()
    message2 = StringVar()
    message3 = StringVar()
    message4 = StringVar()
    message=[message1,message2,message3,message4]
    a = 10
    for i in range(4):
        Entry(add_window, textvariable=message[i],width=23).place(x=150, y=a)      # ввод с клавиатуры
        a+=30

    combo_fak_id = Combobox(add_window)
    combo_fak_id.place(x=150, y=132)
    combo_group_id = Combobox(add_window)
    combo_group_id.place(x=150, y=162)

    combo_fak_id['values'] = list_of_faks          # список факультетов
    combo_fak_id.current(0)

    combo_group_id['values'] = list_of_all_groups       #  список групп
    combo_group_id.current(0)

    Button(add_window, text="    OK    ", command=add_click).place(x=100, y=220) # конструктор BUTTON
    Button(add_window, text="Отмена", command=otmena_click).place(x=165, y=220)  # конструктор BUTTON

    add_window.mainloop()
    return

def change_student():
    list = []
    change_window = Toplevel()                    # создание дополнительного(дочернего) окна
    change_window.title("Изменение инфо студента")  # заголовок окна
    change_window.geometry("500x350+700+500")

    def search():
        global combo_2                            # combo второго списка: выбор id\имени\фамилии
        cur = con.cursor()
        if combo.get() == "ID":
            cur.execute("""SELECT st_id  FROM students """)
        elif combo.get() == "Фамилии":
            cur.execute("""SELECT surname  FROM students """)
        elif combo.get() == "Имени":
            cur.execute("""SELECT name  FROM students """)
        for item in cur:
            list.append(item)                      # список id ИЛИ имен ИЛИ фамилий в зависимости от состояния combo
        Label(change_window, text="Выбор").place(x=10, y=60)
        Button(change_window, text="   OK   ", command=choose).place(x=280, y=55)
        combo_2 = Combobox(change_window)
        combo_2['values'] = list
        combo_2.place(x=90, y=60)
        combo_2.current(0)
        return

    def choose():
        b=120
        global output_item
        change_st=[combo_2.get()]          # значение второго Combobox
        if combo.get() == "ID":
            cur.execute("""SELECT * FROM students WHERE st_id=?""", change_st)
        elif combo.get() == "Фамилии":
            cur.execute("""SELECT * FROM students WHERE surname=?""", change_st)
        elif combo.get() == "Имени":
            cur.execute("""SELECT * FROM students WHERE name=?""", change_st)

        for output_item in cur:
            for i in range(5):        # вывод вертикально:id,имени,фамилии,фак-та и группы выбранного студента
                Label(change_window, text=output_item[i],font=13).place(x=10, y=b)
                b+=50
        Button(change_window, text="Изменить", command=change_id).place(x=90, y=120)
        Button(change_window, text="Изменить", command=change_surname).place(x=90, y=170)
        Button(change_window, text="Изменить", command=change_name).place(x=90, y=220)
        Button(change_window, text="Изменить", command=change_fak).place(x=90, y=270)
        Button(change_window, text="Изменить", command=change_group).place(x=90, y=320)
###########
    def change_id():
        global new_id
        new_id = StringVar()
        Entry(change_window, textvariable=new_id,width=22).place(x=180, y=130)
        Button(change_window, text=" Принять ", command=admit_id).place(x=380, y=120)
    def admit_id():                                # реакция на кнопку "ПРИНЯТЬ"
        list=[new_id.get(),output_item[0]]                # новое и старое ID
        cur.execute("""UPDATE students SET st_id=? WHERE st_id=? """,list)
        con.commit()
###########
    def change_surname():
        global new_surname
        new_surname = StringVar()
        Entry(change_window, textvariable=new_surname,width=22).place(x=180, y=175)
        Button(change_window, text=" Принять ", command=admit_surname).place(x=380, y=170)
    def admit_surname():
        list=[new_surname.get(),output_item[1]]              # новое и старое фамилия
        cur.execute("""UPDATE students SET surname=? WHERE surname=? """,list)
        con.commit()
###########
    def change_name():
        global new_name
        new_name = StringVar()
        Entry(change_window, textvariable=new_name,width=22).place(x=180, y=225)
        Button(change_window, text=" Принять ", command=admit_name).place(x=380, y=220)
    def admit_name():
        list=[new_name.get(),output_item[2]]                   # новое и старое имя
        cur.execute("""UPDATE students SET name=? WHERE name=? """,list)
        con.commit()
###########
    def change_fak():
        global combo_fak_id
        combo_fak_id = Combobox(change_window)
        combo_fak_id.place(x=180, y=270)
        combo_fak_id['values'] = list_of_faks
        combo_fak_id.current(0)
        Button(change_window, text=" Принять ", command=admit_fak).place(x=380, y=270)
        return
    def admit_fak():
        list = [combo_fak_id.get(), output_item[0]]  # новое и старое значение. output_item взят из функции choose()
        cur.execute("""UPDATE students SET fak_id=? WHERE st_id=? """, list)
        if combo_fak_id.get()==list_of_faks[0]:
            new_group=list_of_fak_groups[0]
        elif combo_fak_id.get()==list_of_faks[1]:
            new_group=list_of_e_groups[0]
        elif combo_fak_id.get()==list_of_faks[2]:
            new_group=list_of_m_groups[0]
        list = [new_group,output_item[0]]
        cur.execute("""UPDATE students SET group_id=? WHERE st_id=? """,list)
        con.commit()
        return
###########
    def change_group():
        choosen_groups=[]
        global combo_group_id
        combo_group_id = Combobox(change_window)
        combo_group_id.place(x=180, y=320)
        if output_item[3]==list_of_faks[0]:
            choosen_groups=list_of_fak_groups
        elif output_item[3]==list_of_faks[1]:
            choosen_groups=list_of_e_groups
        elif output_item[3] == list_of_faks[2]:
            choosen_groups = list_of_m_groups

        combo_group_id['values'] = choosen_groups
        combo_group_id.current(0)
        Button(change_window, text=" Принять ", command=admit_group).place(x=380, y=320)
        return
    def admit_group():
        list = [combo_group_id.get(), output_item[0]]
        cur.execute("""UPDATE students SET group_id=? WHERE st_id=? """, list)
        con.commit()
        return

    Label(change_window, text="Найти по").place(x=10, y=20)
    combo = Combobox(change_window)
    combo['values'] = ("ID", "Фамилии", "Имени")             # ищем по id\имени\фамилии
    combo.current(0)                                         # установка варианта( по умолчанию
    combo.place(x=90, y=20)
    btn = Button(change_window, text="   OK   ", command=search).place(x=280, y=15)

    change_window.mainloop()
    return

def change_delete_group():
    change_delete_group_window = Toplevel(m_window)  # создание дополнительного(дочернего) окна
    change_delete_group_window.title("Изменение таблицы GROUPS")  # заголовок окна
    change_delete_group_window.geometry("500x210+700+500")
    def change():
        global new_name
        global old_group_id
        old_group_id=combo.get()
        new_name = StringVar()
        Label(change_delete_group_window, text="Введите новый ID:").place(x=10, y=90)
        Entry(change_delete_group_window,textvariable=new_name,width=22).place(x=150, y=90)
        Button(change_delete_group_window, text=" Принять", command=admit).place(x=335, y=90)
        return
    def delete():
        a=[combo.get()]
        try:
            cur.execute("""DELETE FROM groups WHERE group_id=?""",a)
            con.commit()
        except:
            Label(change_delete_group_window, text="Удаление группы невозможно,т.к."
                                              "в этой группе есть студенты",font=12).place(x=10, y=80)
        return
    def admit():
        list=[new_name.get(),old_group_id]
        cur.execute("""UPDATE groups SET group_id=? WHERE group_id=? """, list)
        con.commit()
        return
    Label(change_delete_group_window, text="Выбор группы").place(x=10, y=20)
    Button(change_delete_group_window, text="Изменить", command=change).place(x=255, y=18)
    Button(change_delete_group_window, text=" Удалить", command=delete).place(x=335, y=18)
    combo = Combobox(change_delete_group_window,width=10)
    combo['values'] = list_of_all_groups
    combo.place(x=130, y=18)
    combo.current(0)

    change_delete_group_window.mainloop()
    return

def delete_student():
    del_window = Toplevel()             # создание дополнительного(дочернего) окна
    del_window.title("Удаление студента")  # заголовок окна
    del_window.geometry("400x250+700+500")
    def otmena_click():               # закрытие окна DELETE STUDENT без сохранения данных
        del_window.after(3, lambda: del_window.destroy())

    def delete():                     # удаление элемента,найденного в SEARCH()
        del_st=[combo_2.get()]          # значение второго Combobox
        if combo.get() == "ID":
            cur.execute("""DELETE FROM students WHERE st_id=?""", del_st)
        elif combo.get() == "Фамилии":
            cur.execute("""DELETE FROM students WHERE surname=?""", del_st)
        elif combo.get() == "Имени":
            cur.execute("""DELETE FROM students WHERE name=?""", del_st)
        con.commit()
        del_window.after(3, lambda: del_window.destroy())

    def search():
        global combo_2
        cur = con.cursor()
        list = []
        if combo.get() == "ID":
            cur.execute("""SELECT st_id  FROM students """)
        elif combo.get() == "Фамилии":
            cur.execute("""SELECT surname  FROM students """)
        elif combo.get() == "Имени":
            cur.execute("""SELECT name  FROM students """)

        for item in cur:
            list.append(item)
        combo_2 = Combobox(del_window)
        combo_2['values'] = list
        combo_2.place(x=120, y=70)
        combo_2.current(0)
        return

    combo = Combobox(del_window)
    combo['values'] = ("ID", "Фамилии", "Имени")
    combo.current(0)                                 # установка варианта по умолчанию
    combo.place(x=120, y=20)

    Button(del_window, text="Найти по", command=search).place(x=30, y=20)
    Button(del_window, text="Удалить студента", command=delete).place(x=30, y=120)
    Button(del_window, text="Отмена", command=otmena_click).place(x=170, y=120)

    del_window.mainloop()
    return

def find_student():
    add_window = Toplevel()                 # создание дополнительного(дочернего) окна
    add_window.title("Поиск студента")      # заголовок окна
    add_window.geometry("450x250+700+500")
    def otmena_click():                     # закрытие окна без сохранения данных
        add_window.after(3, lambda: add_window.destroy())

    def find_click():
        condition=combo.get()
        search = [find_mes.get(),]
        cur = con.cursor()
        if condition=="ID":
            cur.execute("""SELECT st_id ,surname ,name ,fak_id ,group_id ,score FROM students WHERE st_id=?""",search)
        elif condition=="Фамилии":
            cur.execute("""SELECT st_id ,surname ,name ,fak_id ,group_id ,score FROM students WHERE surname=?""", search)
        elif condition =="Имени":
            cur.execute("""SELECT st_id ,surname ,name ,fak_id ,group_id ,score FROM students WHERE name=?""", search)
        count = 0
        for el in cur:
            for i in range(6):
                Label(add_window, text=el,font="Arial 14").place(x=30, y=170)
                count = 1
        if count == 0:
            Label(add_window, text="Такого студента нет в базе",font="Arial 14").place(x=20, y=170)

    Label(add_window, text="Поиск по:").place(x=20, y=20)
    combo = Combobox(add_window)
    combo['values'] = ("ID","Фамилии","Имени" )
    combo.current(0)                                           # установите вариант по умолчанию
    combo.place(x=100,y=20)

    Button(add_window, text="    OK    ", command=find_click).place(x=70, y=120)
    Button(add_window, text="Отмена", command=otmena_click).place(x=140, y=120)

    find_mes = StringVar()
    Entry(add_window, textvariable=find_mes,width=33).place(x=20, y=60)  # ввод с клавиатуры
    add_window.mainloop()
    return
###################################   вывод схемы базы данных
def view_base():
    schema = Toplevel()
    schema.title("Схема базы данных")
    schema.geometry("1000x600+350+100")
    canvas = Canvas(schema,width=1000,height=600)
    canvas.pack()
    img = PhotoImage(file="base_foto.png")
    canvas.create_image(20, 20, anchor=NW, image=img)
    schema.mainloop()
##################################
def main_window():
    m_window.title("Работа с базами данных")   # заголовок окна
    m_window.geometry("850x700+250+1")
    mainmenu = Menu(m_window)
    new_menu = Menu()
    edit_menu= Menu()
    view_menu= Menu()
    mainmenu.add_cascade(label="   FILE   ", menu=new_menu)
    new_menu.add_command(label="Стандартная база студентов", command=choose_standart_tables)
    new_menu.add_command(label="Загрузить базу студентов", command=choose_loaded_tables)
    new_menu.add_separator()
    new_menu.add_command(label="Новый студент",command=new_student)

    mainmenu.add_cascade(label="   EDIT   ",menu=edit_menu)
    edit_menu.add_command(label="Изменить инфо студента",command=change_student)
    edit_menu.add_command(label="Удалить/изменить группу",command=change_delete_group)
    edit_menu.add_separator()
    edit_menu.add_command(label="Удалить студента",command=delete_student)

    mainmenu.add_cascade(label="   VIEW   ",menu=view_menu)
    view_menu.add_command(label="Просмотреть всех студентов  ",command=view_students)
    view_menu.add_command(label="Просмотреть факультеты", command=view_politeh)
    view_menu.add_command(label="Просмотреть группы", command=view_groups)
    view_menu.add_separator()
    view_menu.add_command(label="Найти студента",command=find_student)
    view_menu.add_separator()
    view_menu.add_command(label="Посмотреть схему базы", command=view_base)
    m_window.config(menu=mainmenu)
    m_window.mainloop()
    return

main_window()
con.close()






# main_lbl.place_forget()
# window.after(3000, lambda: window.destroy())  # закрытие окна--3 сек




















